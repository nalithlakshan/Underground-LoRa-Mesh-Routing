import sys
import argparse
import matplotlib.pyplot as plt
sys.path.append('../')
import loraMeshSimulator as sim
import threading
from openpyxl import load_workbook


# Create a lock for thread-safe operations
excel_lock = threading.Lock()

def append_values_to_excel(file_path, sheet_name, values):
    with excel_lock:
        # Load the workbook and select the sheet
        workbook = load_workbook(file_path)
        
        if sheet_name not in workbook.sheetnames:
            # If the sheet does not exist, create it
            workbook.create_sheet(sheet_name)
        
        sheet = workbook[sheet_name]

        # Append values to the end of the sheet
        sheet.append(values)

        # Save the workbook
        workbook.save(file_path)

def main(repeater_delay_multiplier, avg_send_time, total_sim_packets):
    #-------------------------------------------------------------------------------------
    # Simulation config
    #-------------------------------------------------------------------------------------
    node = sim.node
    nodes = sim.nodes
    env =sim.env
    maxDist =sim.maxDist


    sim.avgSendTime = avg_send_time
    sim.repeatDelayMultiplier = repeater_delay_multiplier
    sim.graphics = 0
    sim.realtime_graphics = 0
    sim.debug = 0

    sim.positional_algo = True
    sim.standby_repeater_algo = True
    sim.energy_aware_algo = True

    sim.totalSimPackets = total_sim_packets

    repeaters =[]
    enddevices = []

    #Repeaters
    d = maxDist*0.99
    
    gw1 = sim.node(env, 6.00*d, 1.00*d, "gw")
    repeaters.append(node(env, 5.50*d, 1.00*d, "rp"))
    repeaters.append(node(env, 5.00*d, 1.00*d, "rp"))
    repeaters.append(node(env, 4.50*d, 1.00*d, "rp"))
    repeaters.append(node(env, 4.00*d, 1.00*d, "rp")) #junction 1
    repeaters.append(node(env, 3.50*d, 1.00*d, "rp")) 
    repeaters.append(node(env, 3.00*d, 1.00*d, "rp"))
    repeaters.append(node(env, 2.66*d, 1.00*d, "rp"))
    repeaters.append(node(env, 2.33*d, 1.00*d, "rp")) #junction 2
    repeaters.append(node(env, 2.00*d, 1.00*d, "rp")) 
    repeaters.append(node(env, 1.66*d, 1.00*d, "rp"))

   
    repeaters.append(node(env, 4.00*d, 1.50*d, "rp")) #branch at j1
    repeaters.append(node(env, 4.00*d, 2.00*d, "rp"))
    repeaters.append(node(env, 4.00*d, 2.50*d, "rp"))

    repeaters.append(node(env, 2.33*d, 1.50*d, "rp")) #branch at j2
    repeaters.append(node(env, 2.33*d, 2.00*d, "rp"))
    repeaters.append(node(env, 2.33*d, 2.33*d, "rp"))
    repeaters.append(node(env, 2.33*d, 2.66*d, "rp"))
    gw2 = sim.node(env, 2.33*d, 3.00*d, "gw")
    

    no_of_repeaters = len(repeaters)


    #End devices
    enddevices.append(node(env, 5.50*d, 1.09*d, "ed"))
    enddevices.append(node(env, 5.00*d, 1.09*d, "ed"))
    enddevices.append(node(env, 4.50*d, 1.09*d, "ed"))
    enddevices.append(node(env, 4.00*d, 1.09*d, "ed")) #junction 1
    enddevices.append(node(env, 3.50*d, 1.09*d, "ed")) 
    enddevices.append(node(env, 3.00*d, 1.09*d, "ed"))
    enddevices.append(node(env, 2.66*d, 1.09*d, "ed"))
    enddevices.append(node(env, 2.33*d, 1.09*d, "ed")) #junction 2
    enddevices.append(node(env, 2.00*d, 1.09*d, "ed")) 
    enddevices.append(node(env, 1.66*d, 1.09*d, "ed"))

   
    enddevices.append(node(env, 4.00*d, 1.59*d, "ed")) #branch at j1
    enddevices.append(node(env, 4.00*d, 2.09*d, "ed"))
    enddevices.append(node(env, 4.00*d, 2.59*d, "ed"))

    enddevices.append(node(env, 2.25*d, 1.50*d, "ed")) #branch at j2
    enddevices.append(node(env, 2.25*d, 2.00*d, "ed"))
    enddevices.append(node(env, 2.25*d, 2.33*d, "ed"))
    enddevices.append(node(env, 2.25*d, 2.66*d, "ed"))


    # #Setting distance values of Position Manually
    nodes[0].distanceValue  = 0
    nodes[1].distanceValue  = 40.395
    nodes[2].distanceValue  = 85.577
    nodes[3].distanceValue  = 124.27
    nodes[4].distanceValue  = 170.45
    nodes[5].distanceValue  = 218.14 
    nodes[6].distanceValue  = 199.87
    nodes[7].distanceValue  = 175.31
    nodes[8].distanceValue  = 166.35
    nodes[9].distanceValue  = 175.91 
    nodes[10].distanceValue = 194.42
    nodes[11].distanceValue = 193.34
    nodes[12].distanceValue = 242.02
    nodes[13].distanceValue = 289.93 
    nodes[14].distanceValue = 123.34
    nodes[15].distanceValue = 81.973
    nodes[16].distanceValue = 59.560
    nodes[17].distanceValue = 28.255 
    nodes[18].distanceValue = 0


    # #routing
    nodes[0].nextRp = 0
    nodes[1].nextRp = 0
    nodes[2].nextRp = 0
    nodes[3].nextRp = 1
    nodes[4].nextRp = 2
    nodes[5].nextRp = 3
    nodes[6].nextRp = 14
    nodes[7].nextRp = 14
    nodes[8].nextRp = 15
    nodes[9].nextRp = 14
    nodes[10].nextRp = 14
    nodes[11].nextRp = 3
    nodes[12].nextRp = 4
    nodes[13].nextRp = 11
    nodes[14].nextRp = 16
    nodes[15].nextRp = 18
    nodes[16].nextRp = 18
    nodes[17].nextRp = 18
    nodes[18].nextRp = 18

    
    sim.networkConfig()


    #Sensor Network
    for i in range(len(nodes)):
        if (nodes[i].type.lower() == "ed"):
            env.process(nodes[i].transmit(env))
    #For Energy Aware Mechanism Testing
    # env.process(nodes[28].transm
    

    #prepare show
    if (sim.graphics == 1):
        sim.plt.xlim([0, sim.xmax])
        sim.plt.ylim([0, sim.ymax])
        sim.plt.draw()
        sim.plt.show()

    # start simulation
    env.run()

    #-----------------------------------------------------------------------
    #Print simulation stat
    print ("No of nodes: ", len(nodes)) #FIX
    print ("AvgSendTime (exp. distributed):",sim.avgSendTime)
    print ("Experiment: ", sim.experiment)
    print ("Simulation Time: ",env.now/60000,"mins")
    print ("Full Collision: ", sim.full_collision)
    print ("Air time: ", sim.nodes[1].packet[0].rectime)

    # data extraction rate
    der = len(sim.packetsRecBS)/float(sim.totalSimPackets)
    print("\nOverall DER:", der)

    #Average Latency
    sum_of_latencies = 0
    sim.packetLatencies.sort()
    for i in range(len(sim.packetLatencies)):
        sum_of_latencies += sim.packetLatencies[i]
    average_latency = sum_of_latencies/len(sim.packetLatencies)

    print("Average Latency:",average_latency,"ms")
    print("Minimum Latency:",sim.packetLatencies[0],"ms")
    print("Maximum Latency:",sim.packetLatencies[-1],"ms")

    print("\n Received/Repeated Packets by Each Repeater")
    total_ed_tx_successes = 0
    total_ed_tx_losses = 0
    total_power_consumption = 0
    for i in range(no_of_repeaters):
        rp = repeaters[i]
        ed = enddevices[i]
        
        ed_tx_pkts = len(ed.txPackets)
        ed_tx_successes = len([x for x in ed.txPackets if x in rp.recPackets])
        ed_tx_losses = ed_tx_pkts - ed_tx_successes
        total_ed_tx_successes += ed_tx_successes
        total_ed_tx_losses += ed_tx_losses
        total_power_consumption += rp.batteryCapacity - rp.batteryRemaining
        if(len(ed.txPackets)>0):
            if(len(ed.txPackets) <5):
                print("ED", ed.id, "Sent Pkts:", ed.txPackets)
            print("ED", ed.id, "Sent Pkts:", ed_tx_pkts)
            print("ED", ed.id, "Pkts successfully sent to corresponding repeater:", ed_tx_successes)
            print("ED", ed.id, "Pkts failed to be captured by corresponding repeater:", ed_tx_losses)
            print("ED", ed.id, "percentage of initial Pkt transmission failures:", round(ed_tx_losses/ed_tx_pkts*100,1), "%")
        if(len(rp.recPackets) <20):
            print("RP", rp.id, "Received Pkts:", rp.recPackets)
        print("RP", rp.id, "Received Pkts:", len(rp.recPackets))
        print("RP", rp.id, "Repeated Pkts:", len(rp.txPackets))
        print("RP", rp.id, "percentage time RP was in Tx state:", round(rp.packet[0].rectime*len(rp.txPackets)/env.now*100,1),"%\n")
        

    
    print("********************************************")
    print("Total Generated Packets:", total_sim_packets)
    total_lost_pkts = total_sim_packets - len(set(gw1.recPackets+gw2.recPackets))
    # total_lost_pkts = total_sim_packets - len(set(gw1.recPackets))

    total_intermediate_losses = total_lost_pkts-total_ed_tx_losses
    print("Total Lost Packets:", total_lost_pkts)
    print("---> Lost at initial ED transmission:", total_ed_tx_losses)
    print("---> Lost at intermediate repetition:", total_intermediate_losses, "\n")

    print("\nTotal Standy \t:",sim.total_stanby)
    print("---> Standby Retains \t\t:",sim.standby_retains)
    print("---> Standby Recoveries \t:",sim.standby_recoveries,"\n")
    print("---> Energy Aware Repeater Role Changes \t:",sim.repeater_role_changes,"\n")

    print("\nTotal Power Consumption :", total_power_consumption)

    print("\nOverall DER:", der)

    #print RSSI of packets
    pkt1_tx = 0
    pkt1_rx = 1

    pkt2_tx = 3
    pkt2_rx = 1
    # print("Node",pkt1_tx,"to Node",pkt1_rx,"RSSI:", nodes[pkt1_tx].packet[pkt1_rx].rssi)
    # print("Node",pkt2_tx,"to Node",pkt2_rx,"RSSI:", nodes[pkt2_tx].packet[pkt2_rx].rssi)

    # Append Test to Excel Sheet
    file_path = "representative_network_sim_outputs_Apr2025.xlsx"
    sheet_name = "Sheet1"
    values_to_append = []
    values_to_append.append(sim.experiment)
    values_to_append.append(no_of_repeaters)
    values_to_append.append(repeater_delay_multiplier)
    values_to_append.append(avg_send_time)
    values_to_append.append(total_sim_packets)
    values_to_append.append("---> DER:")
    values_to_append.append(der)
    values_to_append.append("---> Avg Latency:")
    values_to_append.append(average_latency)
    values_to_append.append("---> Min Latency:")
    values_to_append.append(sim.packetLatencies[0])
    values_to_append.append("---> Max Latency:")
    values_to_append.append(sim.packetLatencies[-1])
    # values_to_append.append("---> Total Collisions:")
    # values_to_append.append(len(sim.collidedPackets))
    # values_to_append.append("---> Total Lost Pkts:")
    # values_to_append.append(total_lost_pkts)
    # values_to_append.append("---> Lost at initial ED transmission:")
    # values_to_append.append(total_ed_tx_losses)
    # values_to_append.append("---> Lost at intermediate repetition:")
    # values_to_append.append(total_intermediate_losses)

    append_values_to_excel(file_path, sheet_name, values_to_append)

    # this can be done to keep graphics visible
    if (sim.graphics == 1):
        input('Press Enter to continue ...')


#============================================================
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Repeater Waiting Period Simulation")
    
    # # Add your command-line arguments
    parser.add_argument("-repeater_delay_multiplier" , default=3 , help="How many times the repeater mean waiting period is greater than the pkt transmission air time?")
    parser.add_argument("-avg_send_time"             , default=2000 , help="Average time period of an end-device sending a packet")
    parser.add_argument("-total_sim_packets"         , default=10000 , help="Total number of packets to process in the simulation")

    # # Parse the command-line arguments
    args = parser.parse_args()

    # Call the main function with the parsed arguments
    main(int(args.repeater_delay_multiplier), int(args.avg_send_time), int(args.total_sim_packets))